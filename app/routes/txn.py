from flask import Blueprint, render_template, request, redirect, url_for, flash, session, Response, abort
from flask_login import login_required, current_user
from functools import wraps
from app.models import Transaction, Branch
from app import db
from datetime import datetime
import json, io


txn_bp = Blueprint('txn', __name__)


# ── RBAC Middleware ──────────────────────────────────────────────

def admin_required(f):
    """Decorator: block non-admin users with 403."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            abort(403)
        return f(*args, **kwargs)
    return decorated_function


def _branch_filtered_query():
    """Return a Transaction query pre-filtered by the user's access level.
    Admin: can see everything (optionally filtered by branch_id param).
    Staff: hard-locked to session branch_id only — no param override."""
    query = Transaction.query

    if current_user.is_admin():
        filter_branch_id = request.args.get('branch_id', '')
        if filter_branch_id:
            query = query.filter_by(branch_id=int(filter_branch_id))
    else:
        # Server-side enforcement: staff ALWAYS sees only their branch
        branch_id = session.get('branch_id')
        if not branch_id:
            return None  # will trigger redirect
        query = query.filter_by(branch_id=branch_id)

    return query


# ── Riwayat (History) ────────────────────────────────────────────

@txn_bp.route('/riwayat')
@login_required
def riwayat():
    """Show transaction history with edit/delete options."""
    if 'branch_id' not in session:
        return redirect(url_for('auth.select_branch'))

    filter_type = request.args.get('type', '')
    filter_start_date = request.args.get('start_date', '')
    filter_end_date = request.args.get('end_date', '')
    filter_branch_id = request.args.get('branch_id', '')

    query = _branch_filtered_query()
    if query is None:
        return redirect(url_for('auth.select_branch'))

    if filter_type:
        query = query.filter_by(type=filter_type)

    if filter_start_date:
        try:
            d_start = datetime.strptime(filter_start_date, '%Y-%m-%d').date()
            query = query.filter(Transaction.transaction_date >= d_start)
        except ValueError:
            pass

    if filter_end_date:
        try:
            d_end = datetime.strptime(filter_end_date, '%Y-%m-%d').date()
            query = query.filter(Transaction.transaction_date <= d_end)
        except ValueError:
            pass

    transactions = query.order_by(Transaction.created_at.desc()).limit(200).all()
    all_branches = Branch.query.order_by(Branch.name).all() if current_user.is_admin() else []

    return render_template('riwayat.html',
                           transactions=transactions,
                           filter_type=filter_type,
                           filter_start_date=filter_start_date,
                           filter_end_date=filter_end_date,
                           filter_branch_id=filter_branch_id,
                           all_branches=all_branches)


# ── Edit ─────────────────────────────────────────────────────────

@txn_bp.route('/edit/<int:txn_id>', methods=['GET', 'POST'])
@login_required
def edit(txn_id):
    if 'branch_id' not in session:
        return redirect(url_for('auth.select_branch'))

    txn = Transaction.query.get_or_404(txn_id)

    # Server-side branch check
    if not current_user.is_admin() and txn.branch_id != session['branch_id']:
        abort(403)

    if request.method == 'POST':
        txn.divisi = request.form.get('divisi', txn.divisi)
        txn.nama_bahan = request.form.get('nama_bahan', txn.nama_bahan)
        tanggal = request.form.get('tanggal', '')
        if tanggal:
            try:
                txn.transaction_date = datetime.strptime(tanggal, '%Y-%m-%d').date()
            except ValueError:
                pass

        if txn.type != 'kejadian':
            qty = request.form.get('qty', type=float)
            if qty and qty > 0:
                txn.qty = qty
            txn.satuan = request.form.get('satuan', txn.satuan)

        detail = json.loads(txn.detail_json) if txn.detail_json else {}

        if txn.type == 'masuk':
            detail['asal_barang'] = request.form.get('asal_barang', detail.get('asal_barang', ''))
            detail['keterangan'] = request.form.get('keterangan', '')
        elif txn.type == 'transfer':
            detail['cabang_tujuan'] = request.form.get('cabang_tujuan', detail.get('cabang_tujuan', ''))
            detail['keterangan'] = request.form.get('keterangan', '')
        elif txn.type == 'prepare':
            txn.qty = request.form.get('qty_prepare', 0, type=float)
            txn.satuan = request.form.get('satuan_prepare', txn.satuan)
            detail['qty_prepare'] = request.form.get('qty_prepare', 0, type=float)
            detail['satuan_prepare'] = request.form.get('satuan_prepare', '')
            detail['qty_jadi'] = request.form.get('qty_jadi', 0, type=float)
            detail['satuan_bahan'] = request.form.get('satuan_bahan', '')
            detail['keterangan'] = request.form.get('keterangan', '')
        elif txn.type == 'prepare_in':
            txn.qty = request.form.get('qty_jadi', 0, type=float)
            txn.satuan = request.form.get('satuan_bahan', txn.satuan)
            detail['qty_prepare'] = request.form.get('qty_prepare', 0, type=float)
            detail['satuan_prepare'] = request.form.get('satuan_prepare', '')
            detail['qty_jadi'] = request.form.get('qty_jadi', 0, type=float)
            detail['satuan_bahan'] = request.form.get('satuan_bahan', '')
            detail['keterangan'] = request.form.get('keterangan', '')
        elif txn.type == 'npu':
            detail['keterangan'] = request.form.get('keterangan', '')
        elif txn.type == 'kejadian':
            detail['isi_kejadian'] = request.form.get('isi_kejadian', '')
            txn.nama_bahan = request.form.get('isi_kejadian', '')[:100]

        txn.detail_json = json.dumps(detail)
        db.session.commit()
        flash('Transaksi berhasil diupdate!', 'success')
        return redirect(url_for('txn.riwayat'))

    detail = json.loads(txn.detail_json) if txn.detail_json else {}
    return render_template('edit_txn.html', txn=txn, detail=detail)


# ── Delete ───────────────────────────────────────────────────────

@txn_bp.route('/delete/<int:txn_id>', methods=['POST'])
@login_required
def delete(txn_id):
    txn = Transaction.query.get_or_404(txn_id)

    # Server-side branch check
    if not current_user.is_admin() and txn.branch_id != session.get('branch_id'):
        abort(403)

    db.session.delete(txn)
    db.session.commit()
    flash('Transaksi berhasil dihapus!', 'success')
    return redirect(url_for('txn.riwayat'))


# ── Export Excel (ADMIN ONLY) ────────────────────────────────────

@txn_bp.route('/export')
@login_required
@admin_required
def export_excel():
    """Export transactions as a true Excel (.xlsx) file. Admin only."""
    if 'branch_id' not in session:
        return redirect(url_for('auth.select_branch'))

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    filter_type = request.args.get('type', '')
    filter_start_date = request.args.get('start_date', '')
    filter_end_date = request.args.get('end_date', '')
    filter_branch_id = request.args.get('branch_id', '')

    query = Transaction.query

    if filter_branch_id:
        query = query.filter_by(branch_id=int(filter_branch_id))
        branch_obj = Branch.query.get(int(filter_branch_id))
        branch_name = branch_obj.name if branch_obj else 'Semua_Cabang'
    else:
        branch_name = 'Semua_Cabang'

    if filter_type:
        query = query.filter_by(type=filter_type)

    if filter_start_date:
        try:
            d_start = datetime.strptime(filter_start_date, '%Y-%m-%d').date()
            query = query.filter(Transaction.transaction_date >= d_start)
        except ValueError:
            pass

    if filter_end_date:
        try:
            d_end = datetime.strptime(filter_end_date, '%Y-%m-%d').date()
            query = query.filter(Transaction.transaction_date <= d_end)
        except ValueError:
            pass

    transactions = query.order_by(Transaction.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaksi"

    headers = ['ID', 'Cabang', 'Tanggal', 'Tipe', 'Divisi', 'Nama Bahan', 'Qty', 'Satuan', 'Detail', 'User', 'Dibuat']
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for txn in transactions:
        detail = json.loads(txn.detail_json) if txn.detail_json else {}
        detail_str = '; '.join(f'{k}: {v}' for k, v in detail.items() if v)
        row = [
            txn.id,
            txn.branch.name if txn.branch else '-',
            txn.transaction_date.strftime('%Y-%m-%d') if txn.transaction_date else '-',
            txn.type.upper(),
            txn.divisi,
            txn.nama_bahan,
            txn.qty if txn.type != 'kejadian' else '-',
            txn.satuan if txn.type != 'kejadian' else '-',
            detail_str,
            txn.user.username,
            txn.created_at.strftime('%Y-%m-%d %H:%M')
        ]
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min((max_length + 2), 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    now = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'StockIQ_{branch_name}_{now}.xlsx'

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )
